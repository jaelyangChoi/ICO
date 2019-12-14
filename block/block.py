import difflib
import urllib.request

import hgtk
from bs4 import BeautifulSoup
from konlpy.tag import Okt
from openpyxl import load_workbook, Workbook

from DB.DAO.defaultKeyword import DefaultKeywordDAO
from ml.ml_predict import ModelCombine


class Block:
    def __init__(self):
        self._default_keyword_list = DefaultKeywordDAO().select_all()

    def _searchWord(self, word):
        # 사전 검색 url
        url = "https://stdict.korean.go.kr/api/search.do"
        option = "?certkey_no=1112&key=A8476D8061FAC1B57C5BCE8DA4CDCF28&method=exact"
        query = "&q=" + urllib.parse.quote(word)
        url_query = url + option + query

        # Open API 검색 요청 개체 설정
        request = urllib.request.Request(url_query)

        # 검색 요청 및 처리
        response = urllib.request.urlopen(request)
        rescode = response.getcode()
        if (rescode == 200):
            return response.read().decode('utf-8')
        else:
            return None

    # 국어사전 요청 쿼리함수
    def _wordExistCheck(self, comment):
        # 검색 질의 요청
        res = self._searchWord(comment)
        xmlsoup = BeautifulSoup(res, 'html.parser')
        items = xmlsoup.find_all('item')
        # 단어의 존재 여부 확인
        if len(items) == 0:
            # 단어가 없는 경우
            return False
        else:
            # 단어가 있는 경우
            return True

    # 국어사전 검색함수
    def _tokenize(self, comment):
        print("**문장 내 명사 분리 시작**")
        okt = Okt()
        return [t for t in okt.nouns(comment)]

    # 댓글 품사분리함수(명사만 처리)
    def _stringMatch(self, comment):
        block = 0
        _comment = ""
        default_keywords = self._default_keyword_list

        print("**1차 필터링 시작**")
        print("[" + comment + "]")
        _comment = self._onlyHangul(comment)

        for default_keyword in default_keywords:
            # 차단 키워드 갯수만큼 for문
            if _comment.find(default_keyword.get_keyword()) != -1:
                block = block + 1
                print("매치된 기본 키워드: " + default_keyword.get_keyword())
                break
        #    한글 이외의 것을 제거한 댓글과 키워드 매치
        if block != 0:
            return default_keyword.get_keyword()
        else:
            return "+"

    # String 일치함수, 1차필터링

    def _onlyHangul(self, comment):
        # 특수문자 제거 함수
        _comment = ""
        for j in comment:
            # 댓글 길이만큼 for문
            if hgtk.checker.is_hangul(j):
                _comment += j
            # elif j == ' ':
            #     _comment += j
            #     코멘트 한글자마다 한글인지 파악
            #     한글일 경우 새 String인자에 추가
            else:
                continue
            #      한글이 아니면 추가X
        return _comment

    # 띄어쓰기,특수문자 제외 한글만 추출하는 함수

    def _stringJamoMatch(self, comment):
        _comment = ""
        default_keywords = self._default_keyword_list
        block = 0

        print("**2차 필터링 시작**")
        comment_list = comment
        #   명사 분리된 경우 그냥 하고 아니면 띄어쓰기로 구분

        for j in comment_list:

            _comment = hgtk.text.decompose(j).replace("ᴥ", "").replace(" ", "")

            for default_keyword in default_keywords:

                if default_keyword.get_split_keyword() == None:
                    continue
                #     기본키워드 중 자모음 키워드만 삭제된 경우

                matchRatio = difflib.SequenceMatcher(None, default_keyword.get_split_keyword(), _comment).ratio()

                if matchRatio >= 0.8:
                    # 일치도 75%이상일시 단어가 국어사전에존재하는지 여부 확인, 존재하면 욕X,아니면 욕
                    if self._wordExistCheck(j):
                        print("\t 기본 키워드: " + default_keyword.get_split_keyword())
                        print("\t 존재하는 단어 :" + j + "이므로 차단하지 않습니다")
                        break
                    else:
                        print("기본 키워드: " + default_keyword.get_split_keyword())
                        print("댓글 내 단어: " + _comment)
                        print("일치율: " + str(matchRatio * 100) + "%")
                        block = block + 1
                        break
            if block != 0:
                break
        if block != 0:
            return default_keyword.get_split_keyword() + " & " + _comment + str(matchRatio * 100) + "%"
        else:
            return "+"

    # 유사도판별함수, 2차필터링

    def privateKeywordMatch(self, comments, keywords):

        _comment = ""

        print("**개인 키워드 필터링 시작**")

        for comment in comments:

            block = 0

            if comment['property'] == '-':
                continue
            # 이미 차단된 댓글인 경우 판단하지 않음
            else:
                _comment = self._onlyHangul(comment['comment'])
                # 댓글 내 특수문자 삭제

                for i in keywords:
                    # 기본 키워드 갯수만큼 for문
                    if _comment.find(i) != -1:
                        block = block + 1
                        print("매치된 개인 키워드: " + i)
                        break
                #    한글 이외의 것을 제거한 댓글과 키워드 매치

                if block != 0:

                    comment['property'] = '-'
                    # 차단할 개인 키워드가 있으면 -로 바꿈
                else:
                    continue
                    # 아니면 그대로

        return comments

    # 개인키워드, 3차필터링
    def runBlockComment(self, comment):
        ml = ModelCombine()

        filtering1 = self._stringMatch(comment)
        # 1차 필터링~String일치로 판별

        if filtering1 == "+":
            # 1차 성공시 2차 필터링 시작

            tokenComment = self._tokenize(comment)
            # 품사분리(명사만 추출)
            print(tokenComment)
            filtering2 = self._stringJamoMatch(tokenComment)
            # 자모음 분리 후s 2차 필터링 한번 더

            if filtering2 == "+":

               if str(ml.total_predict(comment)) == '1':
                   return "+"
              # ML도 통과하면 긍정
               else:
                   return "-"

                ####################**********ML로 댓글 넘김***********#############
            else:
                return "-"
                 #  2차에서 부정
        else:
            return "-"
            # 1차에서 부정

    def runBlockCommentInExcel(self):
        ml = ModelCombine()

        load_wb = load_workbook("/Users/77520769/Documents/문해긔/댓글 수집2.xlsx", data_only=True)
        load_ws = load_wb['시트1']
        # 댓글 불러오기
        write_wb = Workbook()
        write_ws = write_wb.active
        # # 저장할 새 엑셀

        for i in range(2, 1178):
            testComment = load_ws['A' + str(i)].value
            write_ws['A' + str(i)] = testComment
            # 새 엑셀에 댓글 저장

            filtering1 = self._stringMatch(testComment)
            # 1차 필터링~String일치로 판별

            if filtering1 == "+":
                # 1차 성공시 2차 필터링 시작
                tokenComment = self._tokenize(testComment)
                # 품사분리(명사만 추출)
                print(tokenComment)
                filtering2 = self._stringJamoMatch(tokenComment)
                # 자모음 분리 후s 2차 필터링 한번 더

                if filtering2 == "+":
                    if str(ml.total_predict(testComment)) == '1':
                        write_ws['B' + str(i)] = 2
                    else:
                        write_ws['B' + str(i)] = 0
                else:
                    write_ws['B' + str(i)] = 0
                    write_ws['C' + str(i)] = filtering2

                # 2차에서 걸리면 중간
            else:
                write_ws['B' + str(i)] = 0
                write_ws['C' + str(i)] = filtering1

        write_wb.save('/Users/77520769/Documents/문해긔/댓글필터링_ML0.7부정_DB수정.xlsx')
